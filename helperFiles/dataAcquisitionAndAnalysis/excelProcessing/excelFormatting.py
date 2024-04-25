# General
import os
import sys
import pandas as pd

# Data interface modules
from openpyxl import load_workbook, Workbook
import pyexcel
import csv

# Aesthetic modules
from openpyxl.styles import Alignment
from openpyxl.styles import Font


class handlingExcelFormat:

    def __init__(self):
        # Hardcoded sheetnames for different types of Excel information
        self.emptySheetName = "empty"
        self.rawSignals_Sheetname = "Raw Signals; File 0"
        self.subjectInfo_SheetName = "Subject Info; File 0"
        self.rawFeatures_AppendedSheetName = " Features; File 0"
        self.experimentalInfo_SheetName = "Experimental Info; File 0"

        # Hardcoded folder names
        self.saveFeatureFolder = "savedFeatures/"

        # Hardcoded file names
        self.saveFeatureFile_Appended = " Features.xlsx"

        # Excel parameters
        self.maxAddToExcelSheet = 1048500  # Max Rows in a Worksheet

    @staticmethod
    def convertToXLSX(inputExcelFile):
        """
        Converts .xls Files to .xlsx Files That OpenPyxl Can Read
        If the File is Already a .xlsx Files, Do Nothing
        If the File is Neither a .xls Nor .xlsx, it Exits the Program
        """
        # Check That the Current Extension is .xls or .xlsx
        _, extension = os.path.splitext(inputExcelFile)
        # If the Extension is .xlsx, the File is Ready; Do Nothing
        if extension == '.xlsx':
            return inputExcelFile
        # If the Extension is Not .xls/.xlsx, Then the Data is in the Wrong Format; Exit Program
        if extension not in ['.xls', '.xlsx']:
            print("Cannot Convert File to .xlsx")
            sys.exit()

        # Create Output File Directory to Save Data ONLY If None Exists
        newExcelFolder = os.path.dirname(inputExcelFile) + "/Excel Files/"
        os.makedirs(newExcelFolder, exist_ok=True)

        # Convert '.xls' to '.xlsx'
        filename = os.path.basename(inputExcelFile)
        newExcelFile = newExcelFolder + filename + "x"
        pyexcel.save_as(file_name=inputExcelFile, dest_file_name=newExcelFile, logfile=open(os.devnull, 'w'))

        # Save New Excel name
        return newExcelFile

    @staticmethod
    def txt2csv(txtFile, csvFile, csvDelimiter=",", overwriteCSV=False):
        # Check to see if csv conversion alreayd happened
        if not os.path.isfile(csvFile) or overwriteCSV:
            with open(txtFile, "r") as inputData:
                in_reader = csv.reader(inputData, delimiter=csvDelimiter)
                with open(csvFile, 'w', newline='') as out_csv:
                    out_writer = csv.writer(out_csv)
                    for row in in_reader:
                        out_writer.writerow(row)

    @staticmethod
    def convertToExcel(inputFile, excelFile, excelDelimiter=",", overwriteXL=False, testSheetNum=0):
        # If the File is Not Already Converted: Convert the CSV to XLSX
        if not os.path.isfile(excelFile) or overwriteXL:
            if excelDelimiter == "fixedWidth":
                df = pd.read_fwf(inputFile)
                df.drop(index=0, inplace=True)  # drop the underlines
                df.to_excel(excelFile, index=False)
                # Load the Data from the Excel File
                xlWorkbook = load_workbook(excelFile, data_only=True, read_only=True)
                xlWorksheets = xlWorkbook.worksheets[testSheetNum:]
            else:
                # Make Excel WorkBook
                xlWorkbook = Workbook()
                xlWorksheet = xlWorkbook.active
                # Write the Data from the CSV File to the Excel WorkBook
                with open(inputFile, "r") as inputData:
                    inReader = csv.reader(inputData, delimiter=excelDelimiter)
                    with open(excelFile, 'w+', newline=''):
                        for row in inReader:
                            xlWorksheet.append(row)
                            # Save as New Excel File
                xlWorkbook.save(excelFile)
                xlWorksheets = [xlWorksheet]
        # Else Load the Data from the Excel File
        else:
            # Load the Data from the Excel File
            xlWorkbook = load_workbook(excelFile, data_only=True, read_only=True)
            xlWorksheets = xlWorkbook.worksheets[testSheetNum:]

        # Return Excel Sheet
        return xlWorkbook, xlWorksheets

    @staticmethod
    def splitExcelSheetsToExcelFiles(inputFile):
        wb = load_workbook(filename=inputFile)

        for sheet in wb.worksheets:
            new_wb = Workbook()
            ws = new_wb.active
            for row_data in sheet.iter_rows():
                for row_cell in row_data:
                    ws[row_cell.coordinate].value = row_cell.value

            new_wb.save('{0}.xlsx'.format(sheet.title))

    @staticmethod
    def addExcelAesthetics(worksheet):
        # Initialize variables
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Loop through each header cell
        for headerCell in worksheet[1]:
            column_cells = worksheet[headerCell.column_letter]

            # Set the column width
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            worksheet.column_dimensions[headerCell.column_letter].width = max(length, worksheet.column_dimensions[headerCell.column_letter].width)
            worksheet.column_dimensions[headerCell.column_letter].bestFit = True
            # Center the Data in the Cells
            for cell in column_cells:
                cell.alignment = align
            # Set the header text color
            headerCell.font = Font(color='00FF0000', italic=True, bold=True)

        return worksheet
