# General Modules
import os
import time
import collections
# Data interface modules
from openpyxl import load_workbook, Workbook

# Import file for handling excel format
from .excelFormatting import handlingExcelFormat


class saveExcelData(handlingExcelFormat):

    def __init__(self):
        super().__init__()

    def getExcelDocument(self, excelFile, overwriteSave=False):
        # If the excel file you are saving already exists.
        if os.path.isfile(excelFile):
            # If You Want to Overwrite the Excel.
            if overwriteSave:
                print("\t\tDeleting Old Excel Workbook")
                os.remove(excelFile)
            else:
                print("\t\tNot overwriting the file ... but your file already exists??")

        # If the File is Not Present: Create The Excel File
        if not os.path.isfile(excelFile):
            print("\t\tCreating New Excel Workbook")
            # Make Excel WorkBook
            WB = Workbook()
            worksheet = WB.active
            worksheet.title = self.emptySheetName
        else:
            print("\t\tExcel File Already Exists. Adding New Sheet to File")
            WB = load_workbook(excelFile, read_only=False)
            worksheet = WB.create_sheet(self.emptySheetName)
        return WB, worksheet

    def addSubjectInfo(self, WB, worksheet, subjectInformationAnswers, subjectInformationQuestions):
        # Assert that the data is in the correct configuration
        assert len(subjectInformationAnswers) == len(subjectInformationQuestions), \
            f"{len(subjectInformationAnswers)} {subjectInformationAnswers} \n {len(subjectInformationQuestions)} {subjectInformationQuestions}"

        # Get the information ready for the file
        header = ["Background Questions", "Answers"]
        subjectInformationPointer = 0

        # Loop through/save all the data in batches of maxAddToExcelSheet.
        for firstIndexInFile in range(0, len(subjectInformationQuestions), self.maxAddToExcelSheet):
            # Add the information to the page
            worksheet.title = self.subjectInfo_SheetName
            worksheet.append(header)  # Add the header labels to this specific file.

            # Add the info to the first page
            while subjectInformationPointer != len(subjectInformationQuestions):
                # Add the data row to the worksheet
                row = [subjectInformationQuestions[subjectInformationPointer], subjectInformationAnswers[subjectInformationPointer]]
                worksheet.append(row)

                subjectInformationPointer += 1
                # Move onto next excel sheet if too much data
                if int(subjectInformationPointer / (firstIndexInFile + 1)) == self.maxAddToExcelSheet:
                    break

            # Finalize document
            worksheet = self.addExcelAesthetics(worksheet)  # Add Excel Aesthetics
            worksheet = WB.create_sheet(self.emptySheetName)  # Add Sheet
        # Remove empty page
        WB.remove(worksheet)

    def addExperimentInfo(self, WB, worksheet, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions):
        # Assert that the data is in the correct configuration
        assert len(experimentTimes) == len(experimentNames)
        assert len(surveyAnswerTimes) == len(surveyAnswersList)
        experimentInfoPointer, featureInfoPointer = 0, 0

        # Get the Header for the experiment and survey
        header = ["Start Experiment (s)", "End Experiment (s)", "Experiment Label", "Feature Collection (s)"]
        header.extend(surveyQuestions)

        # Loop through/save all the data in batches of maxAddToExcelSheet.
        for firstIndexInFile in range(0, max(len(experimentTimes), len(surveyAnswerTimes)), self.maxAddToExcelSheet):
            # Add the information to the page
            worksheet.title = self.experimentalInfo_SheetName
            worksheet.append(header)  # Add the header labels to this specific file.

            # Add the info to the first page
            while experimentInfoPointer != len(experimentTimes) or featureInfoPointer != len(surveyAnswerTimes):
                row = []

                # Add experimental information
                if experimentInfoPointer != len(experimentTimes):
                    row.extend(experimentTimes[experimentInfoPointer])
                    row.append(experimentNames[experimentInfoPointer])
                    experimentInfoPointer += 1
                else:
                    row.extend([None] * 3)
                # Add feature information
                if featureInfoPointer != len(surveyAnswerTimes):
                    row.append(surveyAnswerTimes[featureInfoPointer])
                    row.extend(surveyAnswersList[featureInfoPointer])
                    featureInfoPointer += 1

                # Add the data row to the worksheet
                worksheet.append(row)
                # Move onto the next Excel sheet if too much data
                if int(experimentInfoPointer / (firstIndexInFile + 1)) == self.maxAddToExcelSheet or int(featureInfoPointer / (firstIndexInFile + 1)) == self.maxAddToExcelSheet:
                    break

            # Finalize document
            self.addExcelAesthetics(worksheet)  # Add Excel Aesthetics
            worksheet = WB.create_sheet(self.emptySheetName)  # Add Sheet
        # Remove empty page
        WB.remove(worksheet)

    def saveData(self, deviceType, timepoints, signalData, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList,
                 surveyQuestions, subjectInformationAnswers, subjectInformationQuestions, dataHeaders, saveExcelPath, overwriteSave=False):

        # -------------------- Setup the excel document -------------------- #
        # Create the path to save the Excel file.
        os.makedirs(os.path.dirname(saveExcelPath), exist_ok=True)  # Create Output File Directory to Save Data: If None Exists
        print("\n\tSaving raw signals")

        # Get the excel document.
        WB, worksheet = self.getExcelDocument(saveExcelPath, overwriteSave)

        # -------------- Add experimental/subject information -------------- #

        # Add subject information
        self.addSubjectInfo(WB, worksheet, subjectInformationAnswers, subjectInformationQuestions)
        worksheet = WB.create_sheet(self.emptySheetName)  # Add
        # Add experimental information
        self.addExperimentInfo(WB, worksheet, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions)
        worksheet = WB.create_sheet(self.emptySheetName)  # Add Sheet

        # ---------------------- Add data to document ---------------------- #

        # Create a combined header with time and corresponding signal data columns for each channel
        header = []
        for i, dataHeader in enumerate(dataHeaders):
            header.append(f"Time_{dataHeader.upper()} (s)")
            header.append(f"{dataHeader.upper()} Raw Data")

        # Write the header to the worksheet
        worksheet.title = self.rawSignals_Sheetname
        worksheet.append(header)

        for firstIndexInFile in range(0, len(timepoints), self.maxAddToExcelSheet):
            startTimer = time.time()
            if deviceType == 'empatica':
                # Now append data: align time points with corresponding signal data in the same row
                max_rows = max(len(tp) for tp in timepoints)  # Determine the maximum number of rows to iterate over

                # Iterate through each row (time point)
                for row in range(max_rows):
                    row_data = []  # To hold the row data to append

                    # For each channel (time and data)
                    for channelInd in range(len(timepoints)):
                        if row < len(timepoints[channelInd]):  # Check if the current row exists in the channel data
                            row_data.append(timepoints[channelInd][row])  # Append the time point
                            row_data.append(signalData[channelInd][row])  # Append the corresponding signal data
                        else:
                            # If no data for this row (shorter array), append empty cells
                            row_data.append(None)
                            row_data.append(None)

                    # Append the collected row data to the worksheet
                    worksheet.append(row_data)
            elif deviceType == 'serial':
                max_rows = max(len(tp) for tp in timepoints)  # Determine the maximum number of rows to iterate over

                # Iterate through each row (time point)
                for row in range(max_rows):
                    row_data = []  # To hold the row data to append

                    # For each channel (time and data)
                    for channelInd in range(len(timepoints)):
                        if row < len(timepoints[channelInd]):  # Check if the current row exists in the channel data
                            row_data.append(timepoints[channelInd][row])  # Append the time point
                            row_data.append(signalData[channelInd][row])  # Append the corresponding signal data
                        else:
                            # If no data for this row (shorter array), append empty cells
                            row_data.append(None)
                            row_data.append(None)

                    # Append the collected row data to the worksheet
                    worksheet.append(row_data)

            # Finalize document aesthetics
            self.addExcelAesthetics(worksheet)  # Add Excel Aesthetics
            worksheet = WB.create_sheet(self.emptySheetName)  # Add Sheet

            # Track and estimate time if writing in batches
            maxNumberRows = max(len(tp) for tp in timepoints)
            if firstIndexInFile + self.maxAddToExcelSheet < maxNumberRows:
                endTimer = time.time()
                numberOfSheetsLeft = 1 + (maxNumberRows - firstIndexInFile - self.maxAddToExcelSheet) // self.maxAddToExcelSheet
                timeRemaining = (endTimer - startTimer) * numberOfSheetsLeft
                print("\tEstimated Time Remaining " + str(timeRemaining) + " seconds; Excel Sheets Left to Add: " + str(numberOfSheetsLeft))

        # Remove empty page
        if worksheet.title == self.emptySheetName:
            WB.remove(worksheet)

        # ------------------------------------------------------------------ #
        # ------------------------ Save the document ----------------------- #
        WB.save(saveExcelPath)
        WB.close()
   
    def saveRawFeatures(self, rawFeatureTimesHolder, rawFeatureHolder, biomarkerFeatureNames, biomarkerFeatureOrder, experimentTimes, experimentNames,
                        surveyAnswerTimes, surveyAnswersList, surveyQuestions, subjectInformationAnswers, subjectInformationQuestions, excelFilename, overwriteSave=True):
        print("\n\tSaving raw features")
        # ------------------------------------------------------------------ #
        # -------------------- Setup the excel document -------------------- #
        # Organize variables
        baseFilename = os.path.basename(excelFilename).split('.')[0]
        excelPath = excelFilename.split(baseFilename)[0]

        # Create the path to save the Excel file.
        saveDataFolder = excelPath + self.saveFeatureFolder
        os.makedirs(saveDataFolder, exist_ok=True)  # Create Output File Directory to Save Data: If None Exists
        # Specify the name of the file to save
        saveExcelName = baseFilename + self.saveFeatureFile_Appended
        excelFile = saveDataFolder + saveExcelName

        # Get the excel document.
        WB, worksheet = self.getExcelDocument(excelFile, overwriteSave)

        # ------------------------------------------------------------------ #
        # -------------- Add experimental/subject information -------------- #
        # Add subject information
        self.addSubjectInfo(WB, worksheet, subjectInformationAnswers, subjectInformationQuestions)
        worksheet = WB.create_sheet(self.emptySheetName)  # Add
        # Add experimental information
        self.addExperimentInfo(WB, worksheet, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions)
        worksheet = WB.create_sheet(self.emptySheetName)  # Add Sheet

        # ------------------------------------------------------------------ #
        # ---------------------- Add data to document ---------------------- #  
        counter = collections.Counter()
        biomarkerChannelIndices = [counter.update({item: 1}) or counter[item] - 1 for item in biomarkerFeatureOrder]

        # Individually, add features from each sensor to the excel file.
        for biomarkerInd in range(len(biomarkerFeatureOrder)):
            currentFeatureNames = biomarkerFeatureNames[biomarkerInd]
            channelIndex = biomarkerChannelIndices[biomarkerInd]
            # Extract raw features
            featureTimes = rawFeatureTimesHolder[biomarkerInd]
            rawFeatures = rawFeatureHolder[biomarkerInd]

            # Create the header bar
            header = ["Time (s)"]
            header.extend(currentFeatureNames)

            # Loop through/save all the data in batches of maxAddToExcelSheet.
            for firstIndexInFile in range(0, len(featureTimes), self.maxAddToExcelSheet):
                # Add the information to the page
                worksheet.title = biomarkerFeatureOrder[biomarkerInd].upper() + f" CH{channelIndex}" + self.rawFeatures_AppendedSheetName  # Add the sheet name to the file
                worksheet.append(header)  # Add the header labels to this specific file.

                # Loop through all data to be saved within this sheet in the Excel file.
                for dataInd in range(firstIndexInFile, min(firstIndexInFile + self.maxAddToExcelSheet, len(featureTimes))):
                    # Organize all the data
                    row = [featureTimes[dataInd]]
                    row.extend(rawFeatures[dataInd])

                    # Add the row to the worksheet
                    worksheet.append(row)

                # Finalize document
                self.addExcelAesthetics(worksheet)  # Add Excel Aesthetics
                worksheet = WB.create_sheet(self.emptySheetName)  # Add Sheet
        # Remove empty page
        if worksheet.title == self.emptySheetName:
            WB.remove(worksheet)

        # ------------------------------------------------------------------ #
        # ------------------------ Save the document ----------------------- #  
        # Save as New Excel File
        WB.save(excelFile)
        WB.close()

    def saveFeatureComparison(self, dataMatrix, rowHeaders, colHeaders, saveDataFolder, saveExcelName, sheetName="Feature Comparison", saveFirstSheet=False):
        print("Saving the Data")
        # Create Output File Directory to Save Data: If Not Already Created
        os.makedirs(saveDataFolder, exist_ok=True)

        # Create Path to Save the Excel File
        excelFile = saveDataFolder + saveExcelName
        print(excelFile)

        # If the File is Not Present: Create it
        if not os.path.isfile(excelFile):
            # Make Excel WorkBook
            WB = Workbook()
            WB_worksheet = WB.active
            WB_worksheet.title = sheetName
        else:
            print("Excel File Already Exists. Adding New Sheet to File")
            WB = load_workbook(excelFile)
            WB_worksheet = WB.create_sheet(sheetName)

        maxAddToExcelSheet = 1048500  # Max Rows in a Worksheet
        # Save Data to Worksheet
        for firstIndexInList in range(0, len(dataMatrix), maxAddToExcelSheet):
            # Label First Row
            WB_worksheet.append(colHeaders)

            # Add data to the Worksheet
            for rowInd in range(firstIndexInList, min(firstIndexInList + maxAddToExcelSheet, len(dataMatrix))):
                dataRow = []

                if rowInd < len(rowHeaders):
                    rowHeader = rowHeaders[rowInd]
                    dataRow.append(rowHeader)

                dataRow.extend(dataMatrix[rowInd])
                dataRow[0] = float(dataRow[0])
                dataRow[1] = float(dataRow[1])
                # dataRow[3] = float(dataRow[3])
                # Write the Data to Excel
                WB_worksheet.append(dataRow)

            # Add Excel Aesthetics
            WB_worksheet = self.addExcelAesthetics(WB_worksheet)
            WB_worksheet = WB.create_sheet(sheetName)

            if saveFirstSheet:
                break

        WB.remove(WB_worksheet)
        # Save as New Excel File
        WB.save(excelFile)
        WB.close()
