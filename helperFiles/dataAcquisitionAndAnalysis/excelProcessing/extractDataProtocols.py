# General Modules
import os
import re
import sys
import numpy as np


# Data interface modules
from openpyxl import load_workbook
import csv

# Import file for handling excel format
from .excelFormatting import handlingExcelFormat



class extractData(handlingExcelFormat):

    @staticmethod
    def extractFeatureNames(featureLabelFile, prependedString, appendToName=''):
        """ Extract the Feature Names from a txt File """
        # Check if the File Exists
        if not os.path.exists(featureLabelFile):
            print("The following Input File Does Not Exist:", featureLabelFile)
            sys.exit()

        # Get the Data
        fullText = ''
        with open(featureLabelFile, "r", newline='\n') as inputData:
            inReader = csv.reader(inputData)
            for row in inReader:
                for featureString in row:
                    if featureString[0] != "#":
                        fullText += featureString + ","

        possibleFeatures = fullText.split(prependedString)
        # Extract the Features
        featureList = []
        for feature in possibleFeatures:
            feature = feature.split("[")[-1]
            feature = feature.split("]")[0]
            feature = feature.replace(" ", "")
            feature = feature.replace("\n", "")

            if len(feature) != 0:
                feature = feature.split(",")
                featureList.extend(feature)

        featureListFull = []
        for feature in featureList:
            featureListFull.append(feature + appendToName)

        return featureListFull

    @staticmethod
    def extractRawSignalData(excelSheet, startDataCol=1, endDataCol=2, data=None, streamingOrder=None):
        # Set up parameters.
        timeColumns, dataColumns = [0], []
        dataStartRow = 0

        # If Header Exists, Skip Until You Find the Data
        for row in excelSheet.rows:
            print(row[0].value, row[0].row)
            if type(row[0].value) in [int, float]:
                dataStartRow = row[0].row
                break
            elif "Time" in row[0].value:
                for cellInd in range(1, len(row)):
                    if "Time" in row[cellInd].value: timeColumns.append(cellInd)
                    elif type(row[0].value) is str:
                        if streamingOrder is not None:
                            assert streamingOrder[len(dataColumns)] in row[cellInd].value.lower(), f"streamingOrder: {streamingOrder}; dataColumns: {dataColumns}; row[cellInd].value: {row[cellInd].value}"
                        dataColumns.append(cellInd)
                    else: break

        # Assert the data columns are correct.

        assert endDataCol - startDataCol == dataColumns[-1], f"Data columns: {dataColumns}; timeColumns: {timeColumns}; startDataCol: {startDataCol}; endDataCol: {endDataCol}"
        numFreqs = len(timeColumns)

        # Create the data
        if data is None:
            data = []
            for freqInd in range(numFreqs):
                if len(timeColumns) == freqInd + 1: numBiomarkers = endDataCol - timeColumns[freqInd] - 1
                else: numBiomarkers = timeColumns[freqInd + 1] - timeColumns[freqInd] - 1
                data.append([[], [[] for _ in range(numBiomarkers)]])
        else:
            assert len(data) == numFreqs, f"Data: {data}; numFreqs: {numFreqs}; timeColumns: {timeColumns}"

        # Loop Through the Excel Worksheet to collect all the data
        for dataRow in excelSheet.iter_rows(min_col=startDataCol, min_row=dataStartRow, max_col=endDataCol, max_row=excelSheet.max_row):
            # Stop Collecting Data When there is No More
            if dataRow[0].value is None: break
            freqInd = None

            # Compile the data.
            for columnInd in range(len(dataRow)):
                columnValue = dataRow[columnInd].value
                if columnValue is None: continue

                if columnInd in timeColumns:
                    freqInd = timeColumns.index(columnInd)
                    data[freqInd][0].append(float(columnValue))
                else:
                    lastTimeColumn = timeColumns[freqInd] + 1
                    data[freqInd][1][columnInd - lastTimeColumn].append(float(columnValue or 0))

        return data

    @staticmethod
    def _extractRawEmpaticaData(dataStartRow, excelSheet, startDataCol=1, endDataCol=2, data=None):
        # Initialize the data structure if it's None
        if data is None:
            numPairs = (endDataCol - startDataCol) // 2  # Each pair is a time and data column
            data = [[[] for _ in range(numPairs)], [[] for _ in range(numPairs)]]

        # Now extract data for each time and data pair independently
        for pairInd in range(0, endDataCol - startDataCol, 2):
            timeCol = startDataCol + pairInd
            dataCol = timeCol + 1

            # Loop through the Excel worksheet for the current pair
            for dataRow in excelSheet.iter_rows(min_col=timeCol, min_row=dataStartRow - 1, max_col=dataCol, max_row=excelSheet.max_row):
                timeValue = dataRow[0].value
                dataValue = dataRow[1].value

                # Only add to the list if there is a valid time and data pair
                if timeValue is not None and dataValue is not None:
                    data[0][pairInd // 2].append(float(timeValue))
                    data[1][pairInd // 2].append(float(dataValue))

        return data

    @staticmethod
    def extractExperimentalInfo(excelSheet, experimentTimes=(), experimentNames=(), surveyAnswerTimes=(), surveyAnswersList=(), surveyQuestions=()):
        # If Header Exists, Skip Until You Find the Data
        for row in excelSheet.rows:
            cellA = row[0]
            if type(cellA.value) in [int, float]:
                dataStartRow = cellA.row + 1
                break
            elif type(cellA.value) == str:
                headersFound = []
                for cell in row[4:]:
                    if cell.value is None: break
                    headersFound.append(str(cell.value))
                headersFound = np.asarray(headersFound, dtype=str)
                endDataCol = 4 + len(headersFound)
                # Extract the survey questions if none given
                if len(surveyQuestions) == 0:
                    surveyQuestions = headersFound
                else:
                    # Assert the survey questions are the same in all instances.
                    assert all(surveyQuestions == headersFound), "We have two experimental info sheets with DIFFERENT features"

        # Loop Through the Excel Worksheet to collect all the data
        for dataRow in excelSheet.iter_rows(min_col=1, min_row=dataStartRow - 1, max_col=4, max_row=excelSheet.max_row):
            # Stop Collecting Data When there is No More
            if dataRow[0].value is None:
                break

            # Get Data
            endExperimentTime = float(dataRow[1].value) if dataRow[1].value != None else dataRow[1].value
            experimentTimes.append([float(dataRow[0].value), endExperimentTime])
            experimentNames.append(str(dataRow[2].value))

        # Loop Through the Excel Worksheet to collect all the data
        for dataRow in excelSheet.iter_rows(min_col=4, min_row=dataStartRow - 1, max_col=endDataCol, max_row=excelSheet.max_row):
            # Stop Collecting Data When there is No More
            if dataRow[0].value is None:
                break

            # Get Data
            surveyAnswerTimes.append(float(dataRow[0].value))
            surveyAnswersList.append([float(dataRow[dataInd].value or -1) for dataInd in range(1, len(dataRow))])

        return experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions

    @staticmethod
    def extractSubjectInfo(excelSheet, subjectInformationAnswers=(), subjectInformationQuestions=()):
        dataStartRow = None
        # If Header Exists, Skip Until You Find the Data
        for row in excelSheet.rows:
            cellA = row[0]
            if type(cellA.value) is str:
                dataStartRow = cellA.row + 1
                break

        # Loop Through the Excel Worksheet to collect all the data
        for dataRow in excelSheet.iter_rows(min_col=1, min_row=dataStartRow, max_col=2, max_row=excelSheet.max_row):
            # Stop Collecting Data When there is No More
            if dataRow[0].value is None:
                break

            # Get Data
            subjectInformationAnswers.append(str(dataRow[1].value))
            subjectInformationQuestions.append(str(dataRow[0].value))

        return subjectInformationAnswers, subjectInformationQuestions

    def extractExperimentalData(self, deviceType, worksheets, streamingOrder, surveyQuestions=(), finalSubjectInformationQuestions=()):
        # Initialize data holder
        compiledRawData_eachFreq = None
        numberOfChannels = len(streamingOrder)

        # Initialize experimental information
        experimentTimes, experimentNames = [], []
        surveyAnswerTimes, surveyAnswersList= [], []
        # Initialize subject information
        subjectInformationAnswers, subjectInformationQuestions = [], []

        # Loop through and compile all the data in the file
        for excelSheet in worksheets:
            # Extract experiment information
            if self.experimentalInfo_SheetName in excelSheet.title:
                experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions = self.extractExperimentalInfo(excelSheet, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions)
            # Extract subject information
            elif self.subjectInfo_SheetName in excelSheet.title:
                subjectInformationAnswers, subjectInformationQuestions = self.extractSubjectInfo(excelSheet, subjectInformationAnswers, subjectInformationQuestions)
            # Extract Time and Current Data from the File
            else:
                if deviceType == 'serial': endDataCol = 1 + numberOfChannels
                elif deviceType == 'empatica': endDataCol = 2*numberOfChannels
                else: raise ValueError(f"Unknown device type: {deviceType}")

                # Extract the data
                compiledRawData_eachFreq = self.extractRawSignalData(excelSheet, startDataCol=1, endDataCol=endDataCol, data=compiledRawData_eachFreq, streamingOrder=streamingOrder)

        # Check the data integrity
        if len(compiledRawData_eachFreq[0]) == 0:
            print("\tNo data found in this file")
        # Check that the subject background questions are all the same
        if len(finalSubjectInformationQuestions) != 0:
            assert np.all(np.asarray(finalSubjectInformationQuestions) == subjectInformationQuestions), (
                f"finalSubjectInformationQuestions: {finalSubjectInformationQuestions}; subjectInformationQuestions: {subjectInformationQuestions}")

        return compiledRawData_eachFreq, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions, subjectInformationAnswers, subjectInformationQuestions

    def getData(self, inputFile, deviceType, streamingOrder, testSheetNum=0):
        """
        Extracts Pulse Data from Excel Document (.xlsx). Data can be in any
        worksheet that the user can specify using 'testSheetNum' (0-indexed).
        In the Worksheet:
            Time Data must be in Column 'A' (x-Axis)
            Biolectric Data must be in Column 'B-x' (y-Axis)
        If No Data is present in one cell of a row, it will be read in as zero.
        If deviceType is specified for empatica:
            Time and data axis with different sampling frequency will be saved
            accordingly
            {Major change in the extractExperimentalData section}
        --------------------------------------------------------------------------
        Input Variable Definitions:
            inputFile: The Path to the Excel/TXT/CSV File Containing the Biolectric Data.
            numberOfChannels: The number of biolectric signals to extract.
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) to Begin on.
        --------------------------------------------------------------------------
        """
        # Check if the file exists
        if not os.path.exists(inputFile):
            print("The following Input File Does Not Exist:", inputFile)
            sys.exit()

        # Convert to TXT and CSV Files to XLSX
        if inputFile.endswith(".txt") or inputFile.endswith(".csv"):
            # Extract Filename Information
            oldFileExtension = os.path.basename(inputFile)
            filename = os.path.splitext(oldFileExtension)[0]
            newFilePath = os.path.dirname(inputFile) + "/Excel Files/"
            # Make Output Folder Directory if Not Already Created
            os.makedirs(newFilePath, exist_ok=True)

            # Convert CSV or TXT to XLSX
            excelFile = newFilePath + filename + ".xlsx"
            xlWorkbook, worksheets = self.convertToExcel(inputFile, excelFile, excelDelimiter=",", overwriteXL=False, testSheetNum=testSheetNum)
        # If the File is Already an Excel File, Just Load the File
        elif inputFile.endswith(".xlsx"):
            # Load the Data from the Excel File
            xlWorkbook = load_workbook(inputFile, data_only=True, read_only=True)
            worksheets = xlWorkbook.worksheets[testSheetNum:]
        else:
            raise f"The Following File is Neither CSV, TXT, Nor XLSX: {inputFile}"
        print("Extracting Data from the Excel File:", inputFile)

        # Extract the data
        compiledRawData_eachFreq, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions, subjectInformationAnswers, subjectInformationQuestions = self.extractExperimentalData(deviceType, worksheets, streamingOrder)
        xlWorkbook.close()

        print("\tFinished Collecting Biolectric Data")
        # Finished Data Collection: Close Workbook and Return Data to User
        return compiledRawData_eachFreq, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions, subjectInformationAnswers, subjectInformationQuestions

    @staticmethod
    def extractFeatures(excelSheet, biomarkerFeatureOrder, features, featuresTimesHolder, biomarkerFeatureNames):
        # Assert the integrity of feature extraction
        featureType = excelSheet.title.split(" ")[0].lower()
        assert featureType in biomarkerFeatureOrder, "Please update the biomarkers that we are extracting features from: " + str(featureType)
        # Find the type of features we are extracting    
        channelIndex = int(re.search(pattern=r'CH(\d+)', string=excelSheet.title).group(1)) if " CH" in excelSheet.title else 0
        featureInd = int(np.where(np.asarray(biomarkerFeatureOrder) == featureType)[0][channelIndex])
        if " CH" not in excelSheet.title:
            assert featureInd == biomarkerFeatureOrder.index(featureType), f"Backward compatability broken? -> {excelSheet.title}"

        dataStartRow = None
        # If Header Exists, Skip Until You Find the Data
        for row in excelSheet.rows:
            cellA = row[0]
            if type(cellA.value) in [int, float]:
                dataStartRow = cellA.row + 1
                endDataCol = len(row)
                break
            elif type(cellA.value) == str:
                # If no feature names found, save them
                if len(biomarkerFeatureNames[featureInd]) == 0:
                    biomarkerFeatureNames[featureInd] = np.asarray([str(cell.value) for cell in row[1:]], dtype=str)
                else:
                    # Assert the same feature names present in all files.
                    assert all(biomarkerFeatureNames[featureInd] == np.asarray([str(cell.value) for cell in row[1:]], dtype=str)), "We have two feature sheets with DIFFERENT features for " + featureType + "; " + str(
                        len(np.asarray([str(cell.value) for cell in row[1:]]))) + " " + str(len(biomarkerFeatureNames[featureInd]))

        if dataStartRow is not None:
            # Loop Through the Excel Worksheet to collect all the data
            for dataRow in excelSheet.iter_rows(min_col=1, min_row=dataStartRow - 1, max_col=endDataCol, max_row=excelSheet.max_row):
                # Stop Collecting Data When there is No More
                if dataRow[0].value is None:
                    break

                # Get Data
                featuresTimesHolder[featureInd].append(float(dataRow[0].value))
                features[featureInd].append([float(dataRow[dataInd].value or 0) for dataInd in range(1, len(row))])

        return featuresTimesHolder, features, biomarkerFeatureNames

    def getFeatures(self, biomarkerFeatureOrder, inputFile=None, biomarkerFeatureNames=None, surveyQuestions=(), finalSubjectInformationQuestions=()):
        # Load the Data from the Excel File
        xlWorkbook = load_workbook(inputFile, data_only=True, read_only=True)
        worksheets = xlWorkbook.worksheets

        # Initialize experimental information
        experimentTimes = [];
        experimentNames = []
        surveyAnswerTimes = [];
        surveyAnswersList = [];
        # Initialize subject information
        subjectInformationAnswers = [];
        subjectInformationQuestions = []

        # Initialize data structures for feature parameters.
        featuresHolder = [[] for _ in range(len(biomarkerFeatureOrder))]
        featuresTimesHolder = [[] for _ in range(len(biomarkerFeatureOrder))]
        if biomarkerFeatureNames is None:
            biomarkerFeatureNames = [[] for _ in range(len(biomarkerFeatureOrder))]

        # Loop through and compile all the data in the file
        for excelSheet in worksheets:
            # Extract experiment information
            if self.experimentalInfo_SheetName in excelSheet.title:
                experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions = self.extractExperimentalInfo(excelSheet, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions)
            # Extract subject information
            elif self.subjectInfo_SheetName in excelSheet.title:
                subjectInformationAnswers, subjectInformationQuestions = self.extractSubjectInfo(excelSheet, subjectInformationAnswers, subjectInformationQuestions)
            # Extract the features
            elif self.rawFeatures_AppendedSheetName in excelSheet.title:
                featuresTimesHolder, featuresHolder, biomarkerFeatureNames = self.extractFeatures(excelSheet, biomarkerFeatureOrder, featuresHolder, featuresTimesHolder, biomarkerFeatureNames)
            else: raise f"Unsure what is in this file's excel sheet': {excelSheet.title}"

        # Check that the subject background questions are all the same
        if len(finalSubjectInformationQuestions) != 0:
            assert np.all(np.asarray(finalSubjectInformationQuestions) == subjectInformationQuestions)

        return featuresTimesHolder, featuresHolder, biomarkerFeatureNames, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions, subjectInformationAnswers, subjectInformationQuestions
