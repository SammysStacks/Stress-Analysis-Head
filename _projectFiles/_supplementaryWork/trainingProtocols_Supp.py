# General Modules
import os
import sys

import natsort
import numpy as np
import openpyxl as xl

from ...helperFiles.dataAcquisitionAndAnalysis.excelProcessing.extractDataProtocols import extractData
from ...helperFiles.dataAcquisitionAndAnalysis.excelProcessing.saveDataProtocols import saveExcelData
from ...helperFiles.machineLearning.featureAnalysis.featurePlotting import featurePlotting  # Functions for feature analysis
from ...helperFiles.machineLearning.modelControl.modelSpecifications.compileModelInfo import compileModelInfo


class trainingProtocols(extractData):
    
    def __init__(self, biomarkerFeatureNames, biomarkerOrder, numberOfChannels, trainingDataExcelFolder, readData):
        """
        Parameters
        ----------
        trainingself.trainingDataExcelFolder: The Folder with ONLY the Training Data Excel Files
        """
        super().__init__()
        # General parameters
        self.readData = readData
        self.biomarkerOrder = biomarkerOrder
        self.numberOfChannels = numberOfChannels
        self.biomarkerFeatureNames = biomarkerFeatureNames
        self.trainingDataExcelFolder = trainingDataExcelFolder
        
        # Extract feature information
        self.featureNames = [item for sublist in self.biomarkerFeatureNames for item in sublist]
        
        # Initialize important classes
        self.saveInputs = saveExcelData()
        self.modelInfoClass = compileModelInfo()
        self.analyzeFeatures = featurePlotting(self.trainingDataExcelFolder + "dataAnalysis/", overwrite = False)
    
    def extractData_Old(self, ExcelSheet, startDataCol = 1, endDataCol = 2, data = None):
        # If Header Exists, Skip Until You Find the Data
        for row in ExcelSheet.rows:
            cellA = row[0]
            if type(cellA.value) in [int, float]:
                dataStartRow = cellA.row + 1
                break
        
        if data == None:
            data = [ [], [[] for channel in range(endDataCol-startDataCol)] ]
        # Loop Through the Excel Worksheet to collect all the data
        for dataRow in ExcelSheet.iter_rows(min_col=startDataCol, min_row=dataStartRow-1, max_col=endDataCol, max_row=ExcelSheet.max_row):
            # Stop Collecting Data When there is No More
            if dataRow[0].value == None:
                break
            
            # Get Data
            data[0].append(float(dataRow[0].value))
            for dataInd in range(1, len(dataRow)):
                data[1][dataInd-1].append(float(dataRow[dataInd].value or 0))
                
    def streamBlinkData(self, featureLabelOptions, reanalyzeData = False):
        """
        Parameters
        ----------
        self.trainingDataExcelFolder: The Folder with ONLY the Training Data Excel Files
        """
        # Prepare for data collection
        biomarkerOrder = ["eog"]; subjectOrder = []
        Training_Data = []; Training_Labels = []
                        
        # For each file in the training folder
        for excelFileName in list(natsort.natsorted(os.listdir(self.trainingDataExcelFolder))):
            # Take each excel file
            if excelFileName.endswith(".xlsx") and not excelFileName.startswith("~"):
                # Get Full Path to the Excel File
                excelFile = self.trainingDataExcelFolder + excelFileName
                print("\nLoading Excel File", excelFile)
                
                featureLabel = None
                # Loop through each possible feature label to find correct one
                for possibleFeatureLabelInd in range(len(featureLabelOptions)):
                    possibleFeatureLabel = featureLabelOptions[possibleFeatureLabelInd]
                    
                    # Check if the feature label is found, if so record the final label
                    if possibleFeatureLabel.lower() in excelFile.lower():
                        featureLabel = possibleFeatureLabelInd
                        break
                # If no feature label found, then we messed up
                if featureLabel == None:
                    sys.exit("No Feature Detected in File " + excelFile)
                subjectOrder.append(featureLabel)
                    
                print("\tExtracting Features With Label", featureLabel)                    
                savedFeaturesFile = self.trainingDataExcelFolder + self.saveFeatureFolder + excelFileName.split(".")[0] + self.saveFeatureFile_Appended
                print(savedFeaturesFile)
                # If you want to and can use previously extracted features
                if not reanalyzeData and os.path.isfile(savedFeaturesFile):
                    rawFeatureTimesHolder, rawFeatureHolder, self.biomarkerFeatureNames, experimentTimes, experimentNames, \
                        currentSurveyAnswerTimes, currentSurveyAnswersList, surveyQuestions, currentSubjectInformationAnswers, \
                            subjectInformationQuestions = self.getFeatures(biomarkerOrder, savedFeaturesFile, self.biomarkerFeatureNames, [], [])
                else:
                    # Read in the training file with the raw data
                    print(excelFile)
                    WB = xl.load_workbook(excelFile, data_only=True, read_only=True)
                    worksheets = WB.worksheets
                    
                    compiledRawData = None
                    # Loop through and compile all the data in the file
                    for excelSheet in worksheets:
                        # Compile the data in the sheet
                        compiledRawData = self.extractData_Old(excelSheet, startDataCol = 1, endDataCol = 1 + self.numberOfChannels, data = compiledRawData)
                    # Analyze the data
                    self.readData.resetGlobalVariables()
                    self.readData.streamExcelData(compiledRawData, [], [], [], [], [], [], [], "")
                    # Extract information from the streamed data
                    rawFeatureHolder = self.readData.rawFeatureHolder.copy()
                    rawFeatureTimesHolder = self.readData.rawFeatureTimesHolder.copy()
                    # Remove all previous information from this trial
                    self.readData.resetGlobalVariables()
                        
                    # Save the features to be analyzed in the future.
                    self.saveInputs.saveRawFeatures(rawFeatureTimesHolder, rawFeatureHolder, self.biomarkerFeatureNames, biomarkerOrder, [], [], [],
                                                [], [], [], [], excelFile)


                # Save the features and labels
                Training_Data.extend(rawFeatureHolder.copy())
                Training_Labels.extend(len(rawFeatureHolder.copy())*[featureLabel])
                '''
                import matplotlib.pyplot as plt
                
                plt.plot(compiledRawData[0], compiledRawData[1][0])
                plt.show()
                
                for i in range(len(Training_Data)):
                    print(Training_Labels[i], Training_Data[i])
                    print()
                '''
                

        # Return Training Data and Labels
        return subjectOrder, np.asarray(Training_Data), np.asarray(Training_Labels)
    
    
