#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Need to Install on the Anaconda Prompt:
    $ pip install pyexcel
"""


# -------------------------------------------------------------------------- #
# ---------------------------- Imported Modules ---------------------------- #

# General Modules
import os
import sys
import natsort
import numpy as np
import pandas as pd
# Read/Write to Excel
import csv
import pyexcel
import openpyxl as xl
from openpyxl import load_workbook, Workbook
# Openpyxl Styles
from openpyxl.styles import Alignment
from openpyxl.styles import Font

# -------------------------------------------------------------------------- #
# ---------------------- Extract Test Data from Excel ---------------------- #

class handlingExcelFormat:   

    def __init__(self):
        # Hardcoded sheetnames for different types of excel information
        self.emptySheetName = "empty"
        self.rawSignals_Sheetname = "Raw Signals; File 0"
        self.subjectInfo_SheetName = "Subject Info; File 0"
        self.rawFeatures_AppendedSheetName = " Features; File 0"
        self.experimentalInfo_SheetName = "Experimental Info; File 0"
        
        # Hardcoded folder names
        self.saveFeatureFolder = "savedFeatures/"
        
        # Hardcoded file names
        self.saveFeatureFile_Appended = " Features.xlsx"
        
        # Excel parameters
        self.maxAddToexcelSheet = 1048500  # Max Rows in a Worksheet
        
    def convertToXLSX(self, inputExcelFile):
        """
        Converts .xls Files to .xlsx Files That OpenPyxl Can Read
        If the File is Already a .xlsx Files, Do Nothing
        If the File is Neither a .xls Nor .xlsx, it Exits the Program
        """
        # Check That the Current Extension is .xls or .xlsx
        _, extension = os.path.splitext(inputExcelFile)
        # If the Extension is .xlsx, the File is Ready; Do Nothing
        if extension == '.xlsx':
            return inputExcelFile
        # If the Extension is Not .xls/.xlsx, Then the Data is in the Wrong Format; Exit Program
        if extension not in ['.xls', '.xlsx']:
            print("Cannot Convert File to .xlsx")
            sys.exit()
        
        # Create Output File Directory to Save Data ONLY If None Exists
        newExcelFolder = os.path.dirname(inputExcelFile) + "/Excel Files/"
        os.makedirs(newExcelFolder, exist_ok = True)
        
        # Convert '.xls' to '.xlsx'
        filename = os.path.basename(inputExcelFile)
        newExcelFile = newExcelFolder + filename + "x"
        pyexcel.save_as(file_name = inputExcelFile, dest_file_name = newExcelFile, logfile=open(os.devnull, 'w'))
        
        # Save New Excel name
        return newExcelFile
    
    def txt2csv(self, txtFile, csvFile, csvDelimiter = ",", overwriteCSV = False):
        # Check to see if csv conversion alreayd happened
        if not os.path.isfile(csvFile) or overwriteCSV:
            with open(txtFile, "r") as inputData:
                in_reader = csv.reader(inputData, delimiter = csvDelimiter)
                with open(csvFile, 'w', newline='') as out_csv:
                    out_writer = csv.writer(out_csv)
                    for row in in_reader:
                        out_writer.writerow(row)
    
    def convertToExcel(self, inputFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = 0):
        # If the File is Not Already Converted: Convert the CSV to XLSX
        if not os.path.isfile(excelFile) or overwriteXL:
            if excelDelimiter == "fixedWidth":
                df = pd.read_fwf(inputFile)
                df.drop(index=0, inplace=True) # drop the underlines
                df.to_excel(excelFile, index=False)
                # Load the Data from the Excel File
                xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
                xlWorksheets = xlWorkbook.worksheets[testSheetNum:]
            else:
                # Make Excel WorkBook
                xlWorkbook = xl.Workbook()
                xlWorksheet = xlWorkbook.active
                # Write the Data from the CSV File to the Excel WorkBook
                with open(inputFile, "r") as inputData:
                    inReader = csv.reader(inputData, delimiter = excelDelimiter)
                    with open(excelFile, 'w+', newline=''):
                        for row in inReader:
                            xlWorksheet.append(row)    
                # Save as New Excel File
                xlWorkbook.save(excelFile)
                xlWorksheets = [xlWorksheet]
        # Else Load the Data from the Excel File
        else:
            # Load the Data from the Excel File
            xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
            xlWorksheets = xlWorkbook.worksheets[testSheetNum:]
        
        # Return Excel Sheet
        return xlWorkbook, xlWorksheets
    
    def splitExcelSheetsToExcelFiles(self, inputFile):
        wb = load_workbook(filename=inputFile)
        
        for sheet in wb.worksheets:
            new_wb = Workbook()
            ws = new_wb.active
            for row_data in sheet.iter_rows():
                for row_cell in row_data:
                    ws[row_cell.coordinate].value = row_cell.value
        
            new_wb.save('{0}.xlsx'.format(sheet.title))
    
    def addExcelAesthetics(self, worksheet):
        # Initialize variables
        align = Alignment(horizontal='center',vertical='center',wrap_text=True) 
        
        # Loop through each header cell
        for headerCell in worksheet[1]:
            column_cells = worksheet[headerCell.column_letter]
            
            # Set the column width
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            worksheet.column_dimensions[headerCell.column_letter].width = max(length, worksheet.column_dimensions[headerCell.column_letter].width)
            worksheet.column_dimensions[headerCell.column_letter].bestFit = True
            # Center the Data in the Cells
            for cell in column_cells:
                cell.alignment = align
            # Set the header text color
            headerCell.font = Font(color='00FF0000', italic=True, bold=True)
        
        return worksheet  

class getExcelData(handlingExcelFormat):
    
    def extractData_Old(self, ExcelSheet, startDataCol = 1, endDataCol = 2, data = None):
        # If Header Exists, Skip Until You Find the Data
        for row in ExcelSheet.rows:
            cellA = row[0]
            if type(cellA.value) in [int, float]:
                dataStartRow = cellA.row + 1
                break
        
        if data == None:
            data = [ [], [[] for channel in range(endDataCol-startDataCol)] ]
        # Loop Through the Excel Worksheet to collect all the data
        for dataRow in ExcelSheet.iter_rows(min_col=startDataCol, min_row=dataStartRow-1, max_col=endDataCol, max_row=ExcelSheet.max_row):
            # Stop Collecting Data When there is No More
            if dataRow[0].value == None:
                break
            
            # Get Data
            data[0].append(float(dataRow[0].value))
            for dataInd in range(1, len(dataRow)):
                data[1][dataInd-1].append(float(dataRow[dataInd].value or 0))
        
        return data

    def extractData(self, excelSheet):
        # If Header Exists, Skip Until You Find the Data
        for row in excelSheet.rows:
            cellA = row[0]
            if type(cellA.value) in [int, float]:
                dataStartRow = cellA.row + 1
                break
                
        Voltages = []; serialReads = []
        # Loop Through the Excel Worksheet to collect all the data
        for dataRow in excelSheet.iter_rows(min_col=1, min_row=dataStartRow-1, max_col=2, max_row=excelSheet.max_row):
            # Stop Collecting Data When there is No More
            if dataRow[0].value == None:
                break
            
            # Get Data
            Voltages.append(float(dataRow[0].value))
            serialReads.append(int(dataRow[1].value))

        return Voltages, serialReads
    
    def extractFeatures(self, excelSheet, biomarkerOrder, features, featuresTimesHolder, featureNames):            
        # Find the type of features we are extracting
        featureType = excelSheet.title.split(" ")[0].lower()
        featureInd = biomarkerOrder.index(featureType)
        
        dataStartRow = None
        # If Header Exists, Skip Until You Find the Data
        for row in excelSheet.rows:
            cellA = row[0]
            if type(cellA.value) in [int, float]:
                dataStartRow = cellA.row + 1
                endDataCol = len(row)
                break
            elif type(cellA.value) == str:
                # If no feature names found, save them
                if len(featureNames[featureInd]) == 0:
                    featureNames[featureInd] = np.asarray([str(cell.value) for cell in row[1:]], dtype = str)
                else:
                    # Assert the same feature names present in all files.
                    assert all(featureNames[featureInd] == np.asarray([str(cell.value) for cell in row[1:]], dtype = str)), "We have two feature sheets with DIFFERENT features for " + featureType + "; " + str(len(np.asarray([str(cell.value) for cell in row[1:]]))) + " " + str(len(featureNames[featureInd]))
        
        if dataStartRow != None:
            # Loop Through the Excel Worksheet to collect all the data
            for dataRow in excelSheet.iter_rows(min_col=1, min_row=dataStartRow-1, max_col=endDataCol, max_row=excelSheet.max_row):
                # Stop Collecting Data When there is No More
                if dataRow[0].value == None:
                    break
                
                # Get Data
                featuresTimesHolder[featureInd].append(float(dataRow[0].value))
                features[featureInd].append([float(dataRow[dataInd].value or 0) for dataInd in range(1, len(row))])
                
        return featuresTimesHolder, features, featureNames
    
    
    def getFeatures(self, biomarkerOrder, inputFile = None, featureNames = None, surveyQuestions = [], finalSubjectInformationQuestions = []):
        # Load the Data from the Excel File
        xlWorkbook = xl.load_workbook(inputFile, data_only=True, read_only=True)
        worksheets = xlWorkbook.worksheets
        
        # Initialize experimental information
        experimentTimes = []; experimentNames = []
        surveyAnswerTimes = []; surveyAnswersList = [];
        # Initialize suject information
        subjectInformationAnswers = []; subjectInformationQuestions = []
        
        # Initialize data structures for feature parameters.
        featuresHolder = [[] for _ in range(len(biomarkerOrder))]
        featuresTimesHolder = [[] for _ in range(len(biomarkerOrder))]
        if featureNames == None:
            featureNames = [[] for _ in range(len(biomarkerOrder))]
        
        # Loop through and compile all the data in the file
        for excelSheet in worksheets:
            # Extract experiment information
            if self.experimentalInfo_SheetName in excelSheet.title:
                experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions = self.extractExperimentalInfo(excelSheet, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions)
            # Extract subject information
            elif self.subjectInfo_SheetName in excelSheet.title:
                subjectInformationAnswers, subjectInformationQuestions = self.extractSubjectInfo(excelSheet, subjectInformationAnswers, subjectInformationQuestions)
            # Extract the features
            elif self.rawFeatures_AppendedSheetName in excelSheet.title:
                featuresTimesHolder, featuresHolder, featureNames = self.extractFeatures(excelSheet, biomarkerOrder, featuresHolder, featuresTimesHolder, featureNames)
            else:
                sys.exit("Unsure what is in this file's excel sheet':", excelSheet.title)
    
        # Check that the subject background questions are all the same
        if len(finalSubjectInformationQuestions) != 0:
            assert all(np.asarray(finalSubjectInformationQuestions) == subjectInformationQuestions)
        
        return featuresTimesHolder, featuresHolder, featureNames, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions, subjectInformationAnswers, subjectInformationQuestions
    
    
    def getData(self, inputFile, testSheetNum = 0):
        """
        Extracts Pulse Data from Excel Document (.xlsx). Data can be in any
        worksheet which the user can specify using 'testSheetNum' (0-indexed).
        In the Worksheet:
            Time Data must be in Column 'A' (x-Axis)
            Biolectric Data must be in Column 'B-x' (y-Axis)
        If No Data is present in one cell of a row, it will be read in as zero.
        --------------------------------------------------------------------------
        Input Variable Definitions:
            inputFile: The Path to the Excel/TXT/CSV File Containing the Biolectric Data.
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) to Begin on.
        --------------------------------------------------------------------------
        """
        # Check if File Exists
        if not os.path.exists(inputFile):
            print("The following Input File Does Not Exist:", inputFile)
            sys.exit()
            
        # Convert to TXT and CSV Files to XLSX
        if inputFile.endswith(".txt") or inputFile.endswith(".csv"):
            # Extract Filename Information
            oldFileExtension = os.path.basename(inputFile)
            filename = os.path.splitext(oldFileExtension)[0]
            newFilePath = os.path.dirname(inputFile) + "/Excel Files/"
            # Make Output Folder Directory if Not Already Created
            os.makedirs(newFilePath, exist_ok = True)

            # Convert CSV or TXT to XLSX
            excelFile = newFilePath + filename + ".xlsx"
            xlWorkbook, worksheets = self.convertToExcel(inputFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = testSheetNum)
        # If the File is Already an Excel File, Just Load the File
        elif inputFile.endswith(".xlsx"):
            # Load the Data from the Excel File
            xlWorkbook = xl.load_workbook(inputFile, data_only=True, read_only=True)
            worksheets = xlWorkbook.worksheets
        else:
            print("The Following File is Neither CSV, TXT, Nor XLSX:", inputFile)
        print("Extracting Data from the Excel File:", inputFile)
        
        # Extract the data
        Voltages, serialReads = self.extractData(worksheets[testSheetNum])
        xlWorkbook.close()
        
        # Finished Data Collection: Close Workbook and Return Data to User
        print("\tFinished Collecting Biolectric Data");
        return Voltages, serialReads
    
    def streamBlinkData(self, dataExcelFolder, indivisualFeatureNames, featureLabelOptions, numberOfChannels, readData, reanalyzeData = False):
        """
        Parameters
        ----------
        dataExcelFolder: The Folder with ONLY the Training Data Excel Files
        """
        # Prepare for data collection
        biomarkerOrder = ["eog"]; subjectOrder = []
        Training_Data = []; Training_Labels = []
                        
        # For each file in the training folder
        for excelFileName in list(natsort.natsorted(os.listdir(dataExcelFolder))):
            # Take each excel file
            if excelFileName.endswith(".xlsx") and not excelFileName.startswith("~"):
                # Get Full Path to the Excel File
                excelFile = dataExcelFolder + excelFileName
                print("\nLoading Excel File", excelFile)
                
                featureLabel = None
                # Loop through each possible feature label to find correct one
                for possibleFeatureLabelInd in range(len(featureLabelOptions)):
                    possibleFeatureLabel = featureLabelOptions[possibleFeatureLabelInd]
                    
                    # Check if the feature label is found, if so record the final label
                    if possibleFeatureLabel.lower() in excelFile.lower():
                        featureLabel = possibleFeatureLabelInd
                        break
                # If no feature label found, then we messed up
                if featureLabel == None:
                    sys.exit("No Feature Detected in File " + excelFile)
                subjectOrder.append(featureLabel)
                    
                print("\tExtracting Features With Label", featureLabel)                    
                savedFeaturesFile = dataExcelFolder + self.saveFeatureFolder + excelFileName.split(".")[0] + self.saveFeatureFile_Appended
                print(savedFeaturesFile)
                # If you want to and can use previously extracted features
                if not reanalyzeData and os.path.isfile(savedFeaturesFile):
                    rawFeatureTimesHolder, rawFeatureHolder, indivisualFeatureNames, experimentTimes, experimentNames, currentSurveyAnswerTimes, currentSurveyAnswersList, surveyQuestions, currentSubjectInformationAnswers, subjectInformationQuestions = self.getFeatures(biomarkerOrder, inputFile = savedFeaturesFile, featureNames = indivisualFeatureNames, surveyQuestions = [], finalSubjectInformationQuestions = [])
                else:
                    # Read in the training file with the raw data,
                    WB = xl.load_workbook(excelFile, data_only=True, read_only=True)
                    worksheets = WB.worksheets
                    
                    compiledRawData = None
                    # Loop through and compile all the data in the file
                    for excelSheet in worksheets:
                        # Compile the data in the sheet
                        compiledRawData = self.extractData_Old(excelSheet, startDataCol = 1, endDataCol = 1 + numberOfChannels, data = compiledRawData)
                    # Analyze the data
                    readData.resetGlobalVariables()
                    readData.streamExcelData(compiledRawData, [], [], [], [], [], [], [], "")
                    # Extract information from the streamed data
                    rawFeatureHolder = readData.rawFeatureHolder.copy()
                    rawFeatureTimesHolder = readData.rawFeatureTimesHolder.copy()
                    # Remove all previous information from this trial
                    readData.resetGlobalVariables()
                        
                    # Save the features to be analyzed in the future.
                    saveInputs = saveExcelData()
                    saveInputs.saveRawFeatures(rawFeatureTimesHolder, rawFeatureHolder, indivisualFeatureNames, biomarkerOrder, [], [], [],
                                                [], [], [], [], excelFile)


                # Save the features and labels
                Training_Data.extend(rawFeatureHolder.copy())
                Training_Labels.extend(len(rawFeatureHolder.copy())*[featureLabel])
                '''
                import matplotlib.pyplot as plt
                
                plt.plot(compiledRawData[0], compiledRawData[1][0])
                plt.show()
                
                for i in range(len(Training_Data)):
                    print(Training_Labels[i], Training_Data[i])
                    print()
                '''
                

        # Return Training Data and Labels
        return subjectOrder, np.asarray(Training_Data), np.asarray(Training_Labels)
    
    

class saveExcelData(handlingExcelFormat):
    
    def getExcelDocument(self, excelFile, overwriteSave = False):
        # If the excel file you are saving already exists.
        if os.path.isfile(excelFile):
            # If You Want to Overwrite the Excel.
            if overwriteSave:
                print("\t\tDeleting Old Excel Workbook")
                os.remove(excelFile) 
            else:
                print("\t\tNot overwriting the file ... but your file already exists??")
            
        # If the File is Not Present: Create The Excel File
        if not os.path.isfile(excelFile):
            print("\t\tCreating New Excel Workbook")
            # Make Excel WorkBook
            WB = xl.Workbook()
            worksheet = WB.active 
            worksheet.title = self.emptySheetName
        else:
            print("\t\tExcel File Already Exists. Adding New Sheet to File")
            WB = xl.load_workbook(excelFile, read_only=False)
            worksheet = WB.create_sheet(self.emptySheetName)
        return WB, worksheet
    
    def addExperimentInfo(self, WB, worksheet, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions):
        # Assert that the data is in the correct configuration
        assert len(experimentTimes) == len(experimentNames)
        assert len(surveyAnswerTimes) == len(surveyAnswersList)
        # Set pointer
        experimentInfoPointer = 0; featureInfoPointer = 0
    
        # Get the Header for the experiment and survey
        header = ["Start Experiment (Seconds)", "End Experiment (Seconds)", "Experiment Label"]
        header.append("Feature Collection (Seconds)")
        header.extend(surveyQuestions)
                
        # Loop through/save all the data in batches of maxAddToexcelSheet.
        for firstIndexInFile in range(0, max(len(experimentTimes), len(surveyAnswerTimes)), self.maxAddToexcelSheet):
            # Add the information to the page
            worksheet.title = self.experimentalInfo_SheetName
            worksheet.append(header)  # Add the header labels to this specific file.
            
            # Add the info to the first page
            while experimentInfoPointer != len(experimentTimes) or featureInfoPointer != len(surveyAnswerTimes):
                row = []
                
                # Add experimental information
                if experimentInfoPointer != len(experimentTimes):
                    row.extend(experimentTimes[experimentInfoPointer])
                    row.append(experimentNames[experimentInfoPointer])
                    experimentInfoPointer += 1
                else:
                    row.extend([None]*3)
                # Add feature information
                if featureInfoPointer != len(surveyAnswerTimes):
                    row.append(surveyAnswerTimes[featureInfoPointer])
                    row.extend(surveyAnswersList[featureInfoPointer])
                    featureInfoPointer += 1
                
                # Add the data row to the worksheet
                worksheet.append(row)
                # Move onto next excel sheet if too much data
                if int(experimentInfoPointer/(firstIndexInFile+1)) == self.maxAddToexcelSheet or int(featureInfoPointer/(firstIndexInFile+1)) == self.maxAddToexcelSheet:
                    break

            # Finalize document
            worksheet = self.addExcelAesthetics(worksheet) # Add Excel Aesthetics
            worksheet = WB.create_sheet(self.emptySheetName) # Add Sheet
        # Remove empty page
        WB.remove(worksheet)

    def saveRawFeatures(self, rawFeatureTimesHolder, rawFeatureHolder, indivisualFeatureNames, biomarkerOrder, experimentTimes, experimentNames,
                     surveyAnswerTimes, surveyAnswersList, surveyQuestions, subjectInformationAnswers, subjectInformationQuestions, excelFilename): 
        print("\n\tSaving raw features")
        # ------------------------------------------------------------------ #
        # -------------------- Setup the excel document -------------------- #
        # Organize variables
        baseFilename = os.path.basename(excelFilename).split('.')[0]
        excelPath = excelFilename.split(baseFilename)[0]
        
        # Create the path to save the excel file.
        saveDataFolder = excelPath + self.saveFeatureFolder
        os.makedirs(saveDataFolder, exist_ok=True) # Create Output File Directory to Save Data: If None Exists
        # Specify the name of the file to save
        saveExcelName = baseFilename + self.saveFeatureFile_Appended
        excelFile = saveDataFolder + saveExcelName
        
        # Get the excel document.
        WB, worksheet = self.getExcelDocument(excelFile, overwriteSave = True)
        
        # ------------------------------------------------------------------ #
        # -------------- Add experimental/subject information -------------- #
        # Add subject information
        # self.addSubjectInfo(WB, worksheet, subjectInformationAnswers, subjectInformationQuestions)
        # worksheet = WB.create_sheet(self.emptySheetName) # Add 
        # Add experimental information
        # self.addExperimentInfo(WB, worksheet, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions)
        # worksheet = WB.create_sheet(self.emptySheetName) # Add Sheet

        # ------------------------------------------------------------------ #
        # ---------------------- Add data to document ---------------------- #  
        # Indivisually add features from each sensor to the excel file.
        for featureTypeInd in range(len(biomarkerOrder)):
            featureNames = indivisualFeatureNames[featureTypeInd]
            # Extract raw features
            featureTimes = rawFeatureTimesHolder[featureTypeInd]
            rawFeatures = rawFeatureHolder[featureTypeInd] 
            
            # Create the header bar
            header = ["Time (Seconds)"]
            header.extend(featureNames)
            
            # Loop through/save all the data in batches of maxAddToexcelSheet.
            for firstIndexInFile in range(0, len(featureTimes), self.maxAddToexcelSheet):
                # Add the information to the page
                worksheet.title = biomarkerOrder[featureTypeInd].upper() + self.rawFeatures_AppendedSheetName # Add the sheet name to the file
                worksheet.append(header)  # Add the header labels to this specific file.
                            
                # Loop through all data to be saved within this sheet in the excel file.
                for dataInd in range(firstIndexInFile, min(firstIndexInFile+self.maxAddToexcelSheet, len(featureTimes))):
                    # Organize all the data
                    row = [featureTimes[dataInd]]
                    row.extend(rawFeatures[dataInd])
                                    
                    # Add the row to the worksheet
                    worksheet.append(row)
        
                # Finalize document
                worksheet = self.addExcelAesthetics(worksheet) # Add Excel Aesthetics
                worksheet = WB.create_sheet(self.emptySheetName) # Add Sheet
            
            if len(featureTimes) == 0:
                # Add the information to the page
                worksheet.title = biomarkerOrder[featureTypeInd].upper() + self.rawFeatures_AppendedSheetName # Add the sheet name to the file
                worksheet.append(header)  # Add the header labels to this specific file.
                worksheet = self.addExcelAesthetics(worksheet) # Add Excel Aesthetics
                
        # Remove empty page
        if worksheet.title == self.emptySheetName:
            WB.remove(worksheet)
            
        # ------------------------------------------------------------------ #
        # ------------------------ Save the document ----------------------- #  
        # Save as New Excel File
        WB.save(excelFile)
        WB.close()

    